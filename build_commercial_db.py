#!/usr/bin/env python3
"""
IECC Commercial Secretariat Database Builder
=============================================
Ingests ALL commercial source data into a single SQLite database.

Sources:
  Excel:
    1. 2024_Public_Comments_Commercial_Numbering.xlsx  (CE phase tracking)
    2. 2025_Public_Comments_of_the_2027_Commercial_Numbering.xlsx  (PC phase tracking)
    3. proposals_1767817977.xlsx  (cdpACCESS canonical export)
    4. IECC_PC_Subgroup_Sort_Pass_1.xlsx  (initial subgroup routing)
  JSON:
    5-9. Five subgroup recommendation files (admin, envelope, eplr, hvacr, modeling)
    10.  commercial_consensus_committee_minutes_updated.json
"""

import sqlite3
import json
import re
import os
import hashlib
from datetime import datetime, date
from collections import defaultdict, Counter

import openpyxl

# ─── CONFIG ──────────────────────────────────────────────────────────────────

UPLOADS = "/mnt/user-data/uploads"
PROJECT = "/mnt/project"
DB_PATH = "/home/claude/iecc_commercial.db"

EXCEL_FILES = {
    "ce_commercial": f"{UPLOADS}/2024_Public_Comments_Commercial_Numbering.xlsx",
    "cepc_commercial": f"{UPLOADS}/2025_Public_Comments_of_the_2027_Commercial_Numbering.xlsx",
    "subgroup_sort": f"{UPLOADS}/IECC_PC_Subgroup_Sort_Pass_1.xlsx",
    "cdpaccess": f"{UPLOADS}/proposals_1767817977.xlsx",
}

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def uid(canonical_id):
    """Deterministic UID from canonical ID."""
    return hashlib.sha1(canonical_id.encode()).hexdigest()[:16]

def normalize_id(raw_id):
    """
    Normalize a proposal ID to canonical form.
    Returns (canonical_id, prefix, part, cycle, flags[])
    """
    if not raw_id:
        return None, None, None, None, ["EMPTY_ID"]
    
    raw = str(raw_id).strip()
    flags = []
    part = None
    
    # Extract Part suffix
    part_match = re.search(r'\s+(Part\s+\w+|AM|a)\s*$', raw, re.IGNORECASE)
    if part_match:
        part = part_match.group(1).strip()
        raw = raw[:part_match.start()].strip()
    
    # Handle the garbled HVACR entry
    if raw.startswith('(CEC7)'):
        flags.append("GARBLED_ID_FIXED")
        flags.append("CEC_NORMALIZED_TO_CECP")
        return "CECP7-25", "CEC", part, "2025", flags
    
    # Standard pattern: PREFIX + NUMBER + optional -YEAR
    m = re.match(r'^(CEPC|CECP|CECC|CEC|CE)(\d+)(?:-(\d+))?$', raw)
    if not m:
        flags.append(f"UNPARSEABLE_ID:{raw}")
        return raw, None, part, None, flags
    
    prefix = m.group(1)
    number = m.group(2)
    year = m.group(3)
    
    # Fix missing year suffix
    if not year:
        if prefix in ('CEPC',):
            year = '25'
            flags.append("MISSING_SUFFIX_FIXED")
        elif prefix in ('CEC',):
            year = '25'
            flags.append("MISSING_SUFFIX_FIXED")
        elif prefix == 'CE':
            year = '24'
            flags.append("MISSING_SUFFIX_FIXED")
        else:
            year = '25'
            flags.append("MISSING_SUFFIX_FIXED")
    
    # Normalize CEC -> CECP per procedural rules
    original_prefix = prefix
    if prefix == 'CEC':
        prefix = 'CECP'
        flags.append("CEC_NORMALIZED_TO_CECP")
    
    canonical = f"{prefix}{number}-{year}"
    if part:
        canonical = f"{canonical} {part}"
    
    # Determine cycle
    cycle = "2024" if year in ('24',) else "2025"
    
    return canonical, prefix, part, cycle, flags

def determine_phase(canonical_id, prefix):
    """Determine phase from prefix."""
    if not canonical_id:
        return "UNKNOWN"
    if prefix in ('CE',) or canonical_id.startswith('CE') and not any(canonical_id.startswith(p) for p in ('CEPC','CECP','CECC')):
        # Check it's truly CE and not CEPC/CECP/CECC
        if re.match(r'^CE\d', canonical_id):
            return "CE"
    return "PUBLIC_COMMENT"

def clean_str(val):
    """Clean a cell value to string or None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', 'None', 'none'):
        return None
    return s

def clean_date(val):
    """Convert various date formats to ISO string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s or s == 'None':
        return None
    # Try common formats
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s  # Return as-is if unparseable

def normalize_recommendation(raw):
    """Normalize consensus/subgroup recommendation text."""
    if not raw:
        return None, None, []
    
    s = str(raw).strip()
    flags = []
    
    # Standard forms
    lower = s.lower()
    if lower in ('approved as submitted', 'as submitted'):
        return 'Approved as Submitted', s, flags
    if lower in ('approved as modified', 'as modified', 'as modified 3/25/25'):
        return 'Approved as Modified', s, flags
    if lower in ('approved as further modified',):
        return 'Approved as Modified (Further)', s, flags
    if lower in ('approved',):
        return 'Approved', s, flags
    if lower in ('disapproved',):
        return 'Disapproved', s, flags
    if lower in ('postponed',):
        return 'Postponed', s, flags
    if lower in ('withdrawn', 'wd'):
        return 'Withdrawn', s, flags
    if lower in ('dnp', 'do not process'):
        return 'Do Not Process', s, flags
    if lower in ('remanded',):
        return 'Remanded', s, flags
    if lower in ('referred',):
        return 'Referred', s, flags
    if lower in ('reconsider', 'motion to reconsider'):
        return 'Reconsider', s, flags
    
    # Garbled forms from minutes extraction
    if 'approve' in lower and 'submitted' in lower:
        flags.append(f"GARBLED_REC_FIXED:{s}")
        return 'Approved as Submitted', s, flags
    if 'approve' in lower and 'modified' in lower:
        flags.append(f"GARBLED_REC_FIXED:{s}")
        return 'Approved as Modified', s, flags
    if 'disapprov' in lower:
        flags.append(f"GARBLED_REC_FIXED:{s}")
        return 'Disapproved', s, flags
    if 'remand' in lower:
        flags.append(f"GARBLED_REC_FIXED:{s}")
        return 'Remanded', s, flags
    if lower.startswith('modeling subgroup disapproved'):
        flags.append(f"GARBLED_REC_FIXED:{s}")
        return 'Disapproved', s, flags
    
    # Ballot annotations
    ballot_match = re.search(r'(Ballot \d+:\s*(?:FAILED|PASSED))', s)
    if ballot_match:
        if 'approved' in lower:
            flags.append(f"BALLOT_ANNOTATED:{s}")
            return 'Approved as Submitted', s, flags
        if 'disapprov' in lower:
            flags.append(f"BALLOT_ANNOTATED:{s}")
            return 'Disapproved', s, flags
    
    # Catch-all
    flags.append(f"UNRECOGNIZED_REC:{s}")
    return s, s, flags

def normalize_excel_action(raw):
    """Normalize the consensus action from Excel tracking sheets."""
    if not raw:
        return None
    s = str(raw).strip().lower()
    mapping = {
        'approved': 'Approved',
        'as modified': 'Approved as Modified',
        'as submitted': 'Approved as Submitted',
        'disapproved': 'Disapproved',
        'dnp': 'Do Not Process',
        'do not process': 'Do Not Process',
        'withdrawn': 'Withdrawn',
        'wd': 'Withdrawn',
        'postponed': 'Postponed',
    }
    # Check with stripped lowercase
    for key, val in mapping.items():
        if s == key or s.startswith(key):
            return val
    return str(raw).strip()

# ─── SUBGROUP NAME MAPPING ──────────────────────────────────────────────────

# Maps various subgroup name forms to a single canonical name
SG_CANONICAL = {
    # Excel short forms (2024 tracking)
    'sg1 - electrical power, lighting and renewable': 'Commercial EPLR Subgroup',
    'sg1': 'Commercial EPLR Subgroup',
    'sg2 - envelope and embodied energy': 'Envelope and Embodied Energy Subgroup',
    'sg2': 'Envelope and Embodied Energy Subgroup',
    'sg3- hvacr and water heating': 'Commercial HVACR and Water Heating Subgroup',
    'sg3': 'Commercial HVACR and Water Heating Subgroup',
    'sg4 - modeling, whole building metrics, zero energy': 'Commercial Modeling Subgroup',
    'sg4': 'Commercial Modeling Subgroup',
    'sg6- administrative': 'Commercial Administration Subgroup',
    'sg6': 'Commercial Administration Subgroup',
    # Sort Pass forms
    'commercial administrative': 'Commercial Administration Subgroup',
    'commercial envelope and embodied energy': 'Envelope and Embodied Energy Subgroup',
    'commercial hvac and water heating': 'Commercial HVACR and Water Heating Subgroup',
    'commercial lighting, electrical and renewables': 'Commercial EPLR Subgroup',
    'commercial modeling, whole building metrics and zero energy': 'Commercial Modeling Subgroup',
    # JSON forms (already canonical)
    'commercial administration subgroup': 'Commercial Administration Subgroup',
    'envelope and embodied energy subgroup': 'Envelope and Embodied Energy Subgroup',
    'commercial eplr subgroup': 'Commercial EPLR Subgroup',
    'commercial hvacr and water heating subgroup': 'Commercial HVACR and Water Heating Subgroup',
    'commercial modeling subgroup': 'Commercial Modeling Subgroup',
    # Committee
    'commercial consensus committee': 'Commercial Consensus Committee',
}

def normalize_subgroup(raw):
    if not raw:
        return None
    s = str(raw).strip()
    key = s.lower()
    # Try exact match first
    if key in SG_CANONICAL:
        return SG_CANONICAL[key]
    # Try prefix match
    for k, v in SG_CANONICAL.items():
        if key.startswith(k):
            return v
    return s  # Return as-is if no match

# ─── DATABASE CREATION ───────────────────────────────────────────────────────

def create_db(conn):
    c = conn.cursor()
    c.executescript(f"""
    -- ========================================================================
    -- IECC COMMERCIAL SECRETARIAT DATABASE
    -- Built: {datetime.now().isoformat()}
    -- ========================================================================

    DROP TABLE IF EXISTS data_quality_flags;
    DROP TABLE IF EXISTS consensus_actions;
    DROP TABLE IF EXISTS subgroup_actions;
    DROP TABLE IF EXISTS subgroup_movements;
    DROP TABLE IF EXISTS proposals;
    DROP TABLE IF EXISTS meetings;
    DROP VIEW IF EXISTS v_current_status;
    DROP VIEW IF EXISTS v_ready_for_consensus;
    DROP VIEW IF EXISTS v_data_quality_review;

    CREATE TABLE proposals (
        proposal_uid    TEXT PRIMARY KEY,
        canonical_id    TEXT NOT NULL UNIQUE,
        original_id     TEXT,
        cdpaccess_id    INTEGER,
        cdpaccess_url   TEXT,
        cycle           TEXT NOT NULL,          -- '2024' or '2025'
        phase           TEXT NOT NULL,          -- 'CE' or 'PUBLIC_COMMENT'
        phase_locked    INTEGER DEFAULT 0,
        prefix          TEXT,
        part            TEXT,
        code_section    TEXT,
        proponent       TEXT,
        proponent_email TEXT,
        initial_subgroup TEXT,
        current_subgroup TEXT,
        withdrawn       INTEGER DEFAULT 0,
        withdrawn_date  TEXT,
        withdrawn_reason TEXT,
        source_file     TEXT,
        created_at      TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE subgroup_movements (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        proposal_uid    TEXT NOT NULL REFERENCES proposals(proposal_uid),
        sequence        INTEGER NOT NULL,
        move_from       TEXT,
        move_to         TEXT,
        move_date       TEXT,
        source_file     TEXT,
        UNIQUE(proposal_uid, sequence)
    );

    CREATE TABLE subgroup_actions (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        proposal_uid    TEXT NOT NULL REFERENCES proposals(proposal_uid),
        subgroup        TEXT NOT NULL,
        action_date     TEXT,
        recommendation  TEXT,
        recommendation_raw TEXT,
        vote_for        INTEGER,
        vote_against    INTEGER,
        vote_not_voting INTEGER,
        vote_raw        TEXT,
        reason          TEXT,
        modification_text TEXT,
        notes           TEXT,
        source_file     TEXT,
        UNIQUE(proposal_uid, subgroup)
    );

    CREATE TABLE consensus_actions (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        proposal_uid    TEXT NOT NULL REFERENCES proposals(proposal_uid),
        sequence        INTEGER NOT NULL DEFAULT 1,
        action_date     TEXT,
        recommendation  TEXT,
        recommendation_raw TEXT,
        vote_for        INTEGER,
        vote_against    INTEGER,
        vote_not_voting INTEGER,
        vote_raw        TEXT,
        reason          TEXT,
        modification_text TEXT,
        notes           TEXT,
        moved_by        TEXT,
        seconded_by     TEXT,
        is_final        INTEGER DEFAULT 0,
        source          TEXT,                   -- 'JSON_MINUTES' or 'EXCEL_TRACKING'
        source_file     TEXT,
        UNIQUE(proposal_uid, sequence)
    );

    CREATE TABLE meetings (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        meeting_date    TEXT NOT NULL,
        meeting_time    TEXT,
        body            TEXT DEFAULT 'Commercial Consensus Committee',
        phase           TEXT,
        status          TEXT DEFAULT 'SCHEDULED',
        tentative       INTEGER DEFAULT 0,
        action_count    INTEGER DEFAULT 0,
        notes           TEXT,
        source          TEXT,
        UNIQUE(meeting_date, body)
    );

    CREATE TABLE data_quality_flags (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        proposal_uid    TEXT,
        canonical_id    TEXT,
        table_name      TEXT,
        flag_type       TEXT NOT NULL,
        raw_value       TEXT,
        resolved_value  TEXT,
        needs_review    INTEGER DEFAULT 0,
        created_at      TEXT DEFAULT (datetime('now'))
    );
    """)
    conn.commit()

# ─── INGEST: cdpACCESS EXPORT ────────────────────────────────────────────────

def ingest_cdpaccess(conn):
    """Load canonical proposal data from cdpACCESS export."""
    print("\n[1] Ingesting cdpACCESS export...")
    wb = openpyxl.load_workbook(EXCEL_FILES["cdpaccess"], read_only=True, data_only=True)
    ws = wb["proposals_1767817977"]
    rows = list(ws.iter_rows(values_only=True))
    
    count = 0
    for row in rows[2:]:  # Skip 2 header rows
        cdp_id = row[0]
        agenda = clean_str(row[1])
        if not agenda:
            continue
        # Only commercial
        if not any(agenda.startswith(p) for p in ('CEC', 'CEPC', 'CECP', 'CECC')):
            continue
        
        canonical, orig_prefix, part, cycle, flags = normalize_id(agenda)
        if not canonical:
            continue
        
        phase = determine_phase(canonical, orig_prefix)
        p_uid = uid(canonical)
        
        proponent = clean_str(row[2])
        email = clean_str(row[3])
        sections = clean_str(row[4])
        
        url = f"https://energy.cdpaccess.com/proposal/{cdp_id}/" if cdp_id else None
        
        conn.execute("""
            INSERT OR IGNORE INTO proposals 
            (proposal_uid, canonical_id, original_id, cdpaccess_id, cdpaccess_url,
             cycle, phase, phase_locked, prefix, part, code_section, 
             proponent, proponent_email, source_file)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            p_uid, canonical, agenda, cdp_id, url,
            cycle, phase, 1 if phase == 'CE' else 0,
            orig_prefix, part, sections,
            proponent, email, 'cdpaccess_export'
        ))
        
        for flag in flags:
            conn.execute("""
                INSERT INTO data_quality_flags (proposal_uid, canonical_id, table_name, flag_type, raw_value, resolved_value)
                VALUES (?, ?, 'proposals', ?, ?, ?)
            """, (p_uid, canonical, flag, agenda, canonical))
        
        count += 1
    
    conn.commit()
    wb.close()
    print(f"    → {count} commercial proposals from cdpACCESS")

# ─── INGEST: 2024 COMMERCIAL TRACKING (CE PHASE) ────────────────────────────

def ingest_ce_tracking(conn):
    """Load CE-phase proposals from 2024 Commercial Tracking sheet."""
    print("\n[2] Ingesting 2024 Commercial Tracking (CE phase)...")
    wb = openpyxl.load_workbook(EXCEL_FILES["ce_commercial"], read_only=True, data_only=True)
    ws = wb["Commercial Tracking"]
    rows = list(ws.iter_rows(values_only=True))
    
    proposal_count = 0
    action_count = 0
    
    for row in rows[2:]:  # Skip header rows
        raw_id = clean_str(row[0])
        if not raw_id:
            continue
        if not any(raw_id.startswith(p) for p in ('CE',)):
            continue
        
        canonical, orig_prefix, part, cycle, flags = normalize_id(raw_id)
        if not canonical:
            continue
        
        phase = determine_phase(canonical, orig_prefix)
        p_uid = uid(canonical)
        
        code_section = clean_str(row[1])
        cdp_id = row[2] if row[2] else None
        proponent = clean_str(row[3])
        email = clean_str(row[4])
        subgroup_raw = clean_str(row[5])
        subgroup = normalize_subgroup(subgroup_raw)
        
        url = f"https://energy.cdpaccess.com/proposal/{cdp_id}/" if cdp_id else None
        
        # Insert or update proposal
        conn.execute("""
            INSERT INTO proposals 
            (proposal_uid, canonical_id, original_id, cdpaccess_id, cdpaccess_url,
             cycle, phase, phase_locked, prefix, part, code_section,
             proponent, proponent_email, initial_subgroup, current_subgroup, source_file)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(proposal_uid) DO UPDATE SET
                cdpaccess_id = COALESCE(excluded.cdpaccess_id, proposals.cdpaccess_id),
                cdpaccess_url = COALESCE(excluded.cdpaccess_url, proposals.cdpaccess_url),
                code_section = COALESCE(excluded.code_section, proposals.code_section),
                proponent = COALESCE(excluded.proponent, proposals.proponent),
                proponent_email = COALESCE(excluded.proponent_email, proposals.proponent_email),
                initial_subgroup = COALESCE(excluded.initial_subgroup, proposals.initial_subgroup),
                current_subgroup = COALESCE(excluded.current_subgroup, proposals.current_subgroup)
        """, (
            p_uid, canonical, raw_id, cdp_id, url,
            cycle or '2024', phase, 1 if phase == 'CE' else 0,
            orig_prefix, part, code_section,
            proponent, email, subgroup, subgroup, '2024_commercial_tracking'
        ))
        proposal_count += 1
        
        for flag in flags:
            conn.execute("""
                INSERT INTO data_quality_flags (proposal_uid, canonical_id, table_name, flag_type, raw_value, resolved_value)
                VALUES (?, ?, 'proposals', ?, ?, ?)
            """, (p_uid, canonical, flag, raw_id, canonical))
        
        # Extract movements (columns 6-15: pairs of Move, Date)
        seq = 1
        for i in range(6, 16, 2):
            move_val = clean_str(row[i]) if i < len(row) else None
            date_val = clean_date(row[i+1]) if i+1 < len(row) else None
            if move_val:
                conn.execute("""
                    INSERT OR IGNORE INTO subgroup_movements (proposal_uid, sequence, move_from, move_to, move_date, source_file)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (p_uid, seq, subgroup, move_val, date_val, '2024_commercial_tracking'))
                seq += 1
        
        # Extract consensus action (columns 16-17)
        cons_action = clean_str(row[16]) if len(row) > 16 else None
        cons_date = clean_date(row[17]) if len(row) > 17 else None
        
        if cons_action:
            norm_action = normalize_excel_action(cons_action)
            
            # Handle withdrawn
            if norm_action == 'Withdrawn':
                conn.execute("UPDATE proposals SET withdrawn = 1 WHERE proposal_uid = ?", (p_uid,))
            
            conn.execute("""
                INSERT OR IGNORE INTO consensus_actions 
                (proposal_uid, sequence, action_date, recommendation, recommendation_raw, 
                 is_final, source, source_file)
                VALUES (?, 1, ?, ?, ?, 1, 'EXCEL_TRACKING', '2024_commercial_tracking')
            """, (p_uid, cons_date, norm_action, cons_action))
            action_count += 1
    
    conn.commit()
    wb.close()
    print(f"    → {proposal_count} CE proposals, {action_count} consensus actions")

# ─── INGEST: 2025 COMMERCIAL TRACKING (PUBLIC COMMENT PHASE) ────────────────

def ingest_pc_tracking(conn):
    """Load Public Comment phase proposals from 2025 Commercial Tracking sheet."""
    print("\n[3] Ingesting 2025 Commercial Tracking (Public Comment phase)...")
    wb = openpyxl.load_workbook(EXCEL_FILES["cepc_commercial"], read_only=True, data_only=True)
    ws = wb["Commercial Tracking"]
    rows = list(ws.iter_rows(values_only=True))
    
    proposal_count = 0
    
    for row in rows[2:]:  # Skip header rows
        raw_id = clean_str(row[0])
        if not raw_id:
            continue
        # Only commercial prefixes
        if not any(raw_id.startswith(p) for p in ('CEC', 'CEPC', 'CECP', 'CECC')):
            continue
        
        canonical, orig_prefix, part, cycle, flags = normalize_id(raw_id)
        if not canonical:
            continue
        
        phase = determine_phase(canonical, orig_prefix)
        p_uid = uid(canonical)
        
        code_section = clean_str(row[1])
        cdp_id = row[2] if row[2] else None
        proponent = clean_str(row[3])
        email = clean_str(row[4])
        current_subgroup_raw = clean_str(row[5])
        current_subgroup = normalize_subgroup(current_subgroup_raw)
        initial_raw = clean_str(row[6])
        initial_subgroup = normalize_subgroup(initial_raw)
        
        url = f"https://energy.cdpaccess.com/proposal/{cdp_id}/" if cdp_id else None
        
        conn.execute("""
            INSERT INTO proposals 
            (proposal_uid, canonical_id, original_id, cdpaccess_id, cdpaccess_url,
             cycle, phase, phase_locked, prefix, part, code_section,
             proponent, proponent_email, initial_subgroup, current_subgroup, source_file)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(proposal_uid) DO UPDATE SET
                cdpaccess_id = COALESCE(excluded.cdpaccess_id, proposals.cdpaccess_id),
                cdpaccess_url = COALESCE(excluded.cdpaccess_url, proposals.cdpaccess_url),
                code_section = COALESCE(excluded.code_section, proposals.code_section),
                proponent = COALESCE(excluded.proponent, proposals.proponent),
                proponent_email = COALESCE(excluded.proponent_email, proposals.proponent_email),
                initial_subgroup = COALESCE(excluded.initial_subgroup, proposals.initial_subgroup),
                current_subgroup = COALESCE(excluded.current_subgroup, proposals.current_subgroup)
        """, (
            p_uid, canonical, raw_id, cdp_id, url,
            cycle or '2025', phase, 0,
            orig_prefix, part, code_section,
            proponent, email, initial_subgroup, current_subgroup, '2025_commercial_tracking'
        ))
        proposal_count += 1
        
        for flag in flags:
            conn.execute("""
                INSERT INTO data_quality_flags (proposal_uid, canonical_id, table_name, flag_type, raw_value, resolved_value)
                VALUES (?, ?, 'proposals', ?, ?, ?)
            """, (p_uid, canonical, flag, raw_id, canonical))
        
        # Extract movements (columns 7-24: triplets of MoveFrom, MoveTo, Date)
        seq = 1
        i = 7
        while i + 2 < len(row):
            move_from = clean_str(row[i])
            move_to = clean_str(row[i+1])
            move_date = clean_date(row[i+2])
            if move_from or move_to:
                mf = normalize_subgroup(move_from)
                mt = normalize_subgroup(move_to)
                conn.execute("""
                    INSERT OR IGNORE INTO subgroup_movements (proposal_uid, sequence, move_from, move_to, move_date, source_file)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (p_uid, seq, mf, mt, move_date, '2025_commercial_tracking'))
                seq += 1
            i += 3
    
    conn.commit()
    wb.close()
    print(f"    → {proposal_count} Public Comment proposals")

# ─── INGEST: SUBGROUP SORT PASS 1 ───────────────────────────────────────────

def ingest_sort_pass(conn):
    """Enrich proposals with initial subgroup assignment from Sort Pass 1."""
    print("\n[4] Ingesting Subgroup Sort Pass 1...")
    wb = openpyxl.load_workbook(EXCEL_FILES["subgroup_sort"], read_only=True, data_only=True)
    ws = wb["Assignments"]
    rows = list(ws.iter_rows(values_only=True))
    
    count = 0
    for row in rows[1:]:
        track = clean_str(row[0])
        if track != 'Commercial':
            continue
        
        subgroup_raw = clean_str(row[1])
        raw_id = clean_str(row[2])
        sections = clean_str(row[3])
        
        if not raw_id:
            continue
        
        canonical, orig_prefix, part, cycle, flags = normalize_id(raw_id)
        if not canonical:
            continue
        
        p_uid = uid(canonical)
        subgroup = normalize_subgroup(subgroup_raw)
        
        # Update initial_subgroup if not already set
        conn.execute("""
            UPDATE proposals SET 
                initial_subgroup = COALESCE(proposals.initial_subgroup, ?),
                code_section = COALESCE(proposals.code_section, ?)
            WHERE proposal_uid = ?
        """, (subgroup, sections, p_uid))
        count += 1
    
    conn.commit()
    wb.close()
    print(f"    → {count} commercial assignments enriched")

# ─── INGEST: SUBGROUP RECOMMENDATION JSONs ───────────────────────────────────

def ingest_subgroup_jsons(conn):
    """Load detailed subgroup recommendations from JSON files."""
    print("\n[5] Ingesting subgroup recommendation JSONs...")
    
    json_files = {
        "admin": f"{PROJECT}/admin_subgroup_recommendations_with_dates_v4.json",
        "envelope": f"{PROJECT}/commercial_envelope_subgroup_recommendations_UNIFIED_v2.json",
        "eplr": f"{PROJECT}/commercial_eplr_subgroup_recommendations_UNIFIED.json",
        "hvacr": f"{PROJECT}/commercial_hvacr_subgroup_recommendations.json",
        "modeling": f"{PROJECT}/commercial_modeling_subgroup_recommendations.json",
    }
    
    total = 0
    created = 0
    
    for sg_key, path in json_files.items():
        with open(path) as f:
            data = json.load(f)
        
        records = data.get('records', [])
        file_count = 0
        
        for rec in records:
            raw_id = rec.get('proposal_id', '')
            canonical, orig_prefix, part, cycle, flags = normalize_id(raw_id)
            if not canonical:
                continue
            
            phase = determine_phase(canonical, orig_prefix)
            p_uid = uid(canonical)
            
            subgroup = normalize_subgroup(rec.get('subgroup', ''))
            
            # Ensure proposal exists
            existing = conn.execute("SELECT 1 FROM proposals WHERE proposal_uid = ?", (p_uid,)).fetchone()
            if not existing:
                conn.execute("""
                    INSERT INTO proposals 
                    (proposal_uid, canonical_id, original_id, cycle, phase, phase_locked, 
                     prefix, part, code_section, proponent, initial_subgroup, current_subgroup, source_file)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    p_uid, canonical, raw_id,
                    cycle or ('2024' if phase == 'CE' else '2025'),
                    phase, 1 if phase == 'CE' else 0,
                    orig_prefix, part,
                    rec.get('code_section'), rec.get('proponent'),
                    subgroup, subgroup, f'json_{sg_key}'
                ))
                created += 1
            
            for flag in flags:
                conn.execute("""
                    INSERT INTO data_quality_flags (proposal_uid, canonical_id, table_name, flag_type, raw_value, resolved_value)
                    VALUES (?, ?, 'subgroup_actions', ?, ?, ?)
                """, (p_uid, canonical, flag, raw_id, canonical))
            
            # Normalize recommendation
            rec_clean, rec_raw_out, rec_flags = normalize_recommendation(rec.get('recommendation'))
            for flag in rec_flags:
                conn.execute("""
                    INSERT INTO data_quality_flags (proposal_uid, canonical_id, table_name, flag_type, raw_value, resolved_value)
                    VALUES (?, ?, 'subgroup_actions', ?, ?, ?)
                """, (p_uid, canonical, flag, rec.get('recommendation'), rec_clean))
            
            conn.execute("""
                INSERT OR REPLACE INTO subgroup_actions 
                (proposal_uid, subgroup, action_date, recommendation, recommendation_raw,
                 vote_for, vote_against, vote_not_voting, vote_raw,
                 reason, modification_text, notes, source_file)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                p_uid, subgroup, rec.get('action_date'),
                rec_clean, rec.get('recommendation'),
                rec.get('vote_for'), rec.get('vote_against'), rec.get('vote_not_voting'),
                rec.get('vote_raw'),
                rec.get('reason'), rec.get('modification_text'), rec.get('notes'),
                f'json_{sg_key}'
            ))
            file_count += 1
        
        total += file_count
        print(f"    {sg_key}: {file_count} records")
    
    conn.commit()
    print(f"    → {total} subgroup actions total ({created} new proposals created)")

# ─── INGEST: CONSENSUS MINUTES JSON ─────────────────────────────────────────

def ingest_consensus_json(conn):
    """Load detailed consensus committee actions from minutes JSON."""
    print("\n[6] Ingesting consensus committee minutes JSON...")
    
    path = f"{PROJECT}/commercial_consensus_committee_minutes_updated.json"
    with open(path) as f:
        data = json.load(f)
    
    records = data.get('records', [])
    
    # Group by proposal to establish sequence
    by_proposal = defaultdict(list)
    for rec in records:
        raw_id = rec.get('proposal_id', '')
        canonical, orig_prefix, part, cycle, flags = normalize_id(raw_id)
        if canonical:
            key = canonical
            by_proposal[key].append((canonical, orig_prefix, part, cycle, flags, rec))
    
    # Track meetings
    meeting_dates = Counter()
    
    total = 0
    created = 0
    dupes_skipped = 0
    
    for key, entries in by_proposal.items():
        # Sort by date to establish sequence
        entries.sort(key=lambda x: x[5].get('action_date', ''))
        
        # Detect exact duplicates
        seen = set()
        deduped = []
        for entry in entries:
            rec = entry[5]
            fingerprint = (rec.get('action_date'), rec.get('recommendation'), 
                          rec.get('vote_for'), rec.get('vote_against'))
            if fingerprint in seen:
                dupes_skipped += 1
                canonical = entry[0]
                p_uid_temp = uid(canonical)
                conn.execute("""
                    INSERT INTO data_quality_flags (proposal_uid, canonical_id, table_name, flag_type, raw_value)
                    VALUES (?, ?, 'consensus_actions', 'EXACT_DUPLICATE_SKIPPED', ?)
                """, (p_uid_temp, canonical, f"{rec.get('action_date')}|{rec.get('recommendation')}"))
                continue
            seen.add(fingerprint)
            deduped.append(entry)
        
        for seq_idx, (canonical, orig_prefix, part, cycle, flags, rec) in enumerate(deduped, 1):
            phase = determine_phase(canonical, orig_prefix)
            p_uid = uid(canonical)
            
            # Ensure proposal exists
            existing = conn.execute("SELECT 1 FROM proposals WHERE proposal_uid = ?", (p_uid,)).fetchone()
            if not existing:
                conn.execute("""
                    INSERT INTO proposals 
                    (proposal_uid, canonical_id, original_id, cycle, phase, phase_locked, 
                     prefix, part, source_file)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    p_uid, canonical, rec.get('proposal_id'),
                    cycle or ('2024' if phase == 'CE' else '2025'),
                    phase, 1 if phase == 'CE' else 0,
                    orig_prefix, part, 'json_consensus'
                ))
                created += 1
            
            for flag in flags:
                conn.execute("""
                    INSERT INTO data_quality_flags (proposal_uid, canonical_id, table_name, flag_type, raw_value, resolved_value)
                    VALUES (?, ?, 'consensus_actions', ?, ?, ?)
                """, (p_uid, canonical, flag, rec.get('proposal_id'), canonical))
            
            # Normalize recommendation
            rec_clean, rec_raw_out, rec_flags = normalize_recommendation(rec.get('recommendation'))
            for flag in rec_flags:
                conn.execute("""
                    INSERT INTO data_quality_flags (proposal_uid, canonical_id, table_name, flag_type, raw_value, resolved_value)
                    VALUES (?, ?, 'consensus_actions', ?, ?, ?)
                """, (p_uid, canonical, flag, rec.get('recommendation'), rec_clean))
            
            is_final = 1 if seq_idx == len(deduped) else 0
            
            # Parse moved_by / seconded_by from notes
            notes = rec.get('notes', '') or ''
            moved_by = None
            seconded_by = None
            m = re.search(r'Moved by:\s*(.+?)(?:;|$)', notes)
            if m:
                moved_by = m.group(1).strip()
            m = re.search(r'Seconded by:\s*(.+?)(?:;|$)', notes)
            if m:
                seconded_by = m.group(1).strip()
            
            meeting_dates[rec.get('action_date')] += 1
            
            conn.execute("""
                INSERT OR IGNORE INTO consensus_actions 
                (proposal_uid, sequence, action_date, recommendation, recommendation_raw,
                 vote_for, vote_against, vote_not_voting, vote_raw,
                 reason, modification_text, notes, moved_by, seconded_by,
                 is_final, source, source_file)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'JSON_MINUTES', 'json_consensus')
            """, (
                p_uid, seq_idx, rec.get('action_date'),
                rec_clean, rec.get('recommendation'),
                rec.get('vote_for'), rec.get('vote_against'), rec.get('vote_not_voting'),
                rec.get('vote_raw'),
                rec.get('reason'), rec.get('modification_text'), rec.get('notes'),
                moved_by, seconded_by, is_final
            ))
            total += 1
    
    # Upsert consensus actions from JSON over Excel where both exist
    # The JSON data has more detail (votes, reasons, movers) so it takes precedence
    # Mark JSON consensus actions as final where they override Excel
    conn.execute("""
        UPDATE consensus_actions SET is_final = 0 
        WHERE source = 'EXCEL_TRACKING' 
        AND proposal_uid IN (SELECT proposal_uid FROM consensus_actions WHERE source = 'JSON_MINUTES')
    """)
    
    # Insert meetings
    for meeting_date, count in sorted(meeting_dates.items()):
        phase = 'CE' if meeting_date and meeting_date < '2026-01-01' else 'PUBLIC_COMMENT'
        conn.execute("""
            INSERT INTO meetings (meeting_date, body, phase, status, action_count, source) 
            VALUES (?, 'Commercial Consensus Committee', ?, 'COMPLETED', ?, 'JSON_MINUTES')
            ON CONFLICT(meeting_date, body) DO UPDATE SET 
                status = 'COMPLETED', action_count = excluded.action_count, phase = excluded.phase
        """, (meeting_date, phase, count))
    
    conn.commit()
    print(f"    → {total} consensus actions ({created} new proposals, {dupes_skipped} duplicates skipped)")
    print(f"    → {len(meeting_dates)} meetings recorded")

# ─── INGEST: MEETING SCHEDULE (from PDF) ─────────────────────────────────────

def ingest_schedule(conn):
    """
    Ingest the official meeting schedule from 260221 PDF.
    Parsed from OCR text — the PDF is a grid layout.
    Schedule source: 260221_Consensus_Committees_and_Subgroups_Meeting_Schedules.pdf
    Modified 02/21/26
    """
    print("\n[SCHEDULE] Ingesting meeting schedule from 260221 PDF...")
    
    # ── COMMERCIAL TRACK ──
    # All times Eastern unless noted
    
    schedule = [
        # Commercial Consensus Committee — Wednesday pm
        # "Wednesday pm (2/25, 3/4, 3/11, 3/25)" + tentative 4/22, 4/29
        ("2026-02-25", "2:00 p.m. to 5:00 p.m. EST", "Commercial Consensus Committee", 0),
        ("2026-03-04", "2:00 p.m. to 5:00 p.m. EST", "Commercial Consensus Committee", 0),
        ("2026-03-11", "2:00 p.m. to 5:00 p.m. EST", "Commercial Consensus Committee", 0),
        ("2026-03-25", "2:00 p.m. to 5:00 p.m. EST", "Commercial Consensus Committee", 0),
        ("2026-04-22", "2:00 p.m. to 5:00 p.m. EST", "Commercial Consensus Committee", 1),
        ("2026-04-29", "2:00 p.m. to 5:00 p.m. EST", "Commercial Consensus Committee", 1),
        
        # CO Administrative — "Wednesday pm (2/25, 3/4, 3/11, 3/25)" wait no...
        # From OCR: CO Administrative line shows: "Wednesday am (2/25 from 11 a.m. to 2 p.m. EST) (3/3, 3/12, 3/26, 4/16)"
        #   and tentative: 4/7, 4/21
        # Actually re-reading: body mapping from OCR columns:
        # CO Administrative: Tuesday pm (3/3 from 11:00 a.m. to 2:00 p.m. EST) (3/12, 3/26, 4/16)
        #   tentative: 4/7, 4/21
        ("2026-03-03", "11:00 a.m. to 2:00 p.m. EST", "Commercial Administration Subgroup", 0),
        ("2026-03-12", "2:00 p.m. EST", "Commercial Administration Subgroup", 0),
        ("2026-03-26", "2:00 p.m. EST", "Commercial Administration Subgroup", 0),
        ("2026-04-16", "2:00 p.m. EST", "Commercial Administration Subgroup", 0),
        ("2026-04-07", "2:00 p.m. EST", "Commercial Administration Subgroup", 1),
        ("2026-04-21", "2:00 p.m. EST", "Commercial Administration Subgroup", 1),
        
        # CO EPLR — "Wednesday am (3/4, 3/18, 4/15 - starting at 10:00 a.m. EST)"
        #   tentative: 4/13, 4/27
        ("2026-03-04", "10:00 a.m. EST", "Commercial EPLR Subgroup", 0),
        ("2026-03-18", "10:00 a.m. EST", "Commercial EPLR Subgroup", 0),
        ("2026-04-15", "10:00 a.m. EST", "Commercial EPLR Subgroup", 0),
        ("2026-04-13", "10:00 a.m. EST", "Commercial EPLR Subgroup", 1),
        ("2026-04-27", "10:00 a.m. EST", "Commercial EPLR Subgroup", 1),
        
        # CO Envelope — "Thursday pm (3/5, 3/19)"
        #   tentative: 4/16, 4/30
        ("2026-03-05", "2:00 p.m. EST", "Envelope and Embodied Energy Subgroup", 0),
        ("2026-03-19", "2:00 p.m. EST", "Envelope and Embodied Energy Subgroup", 0),
        ("2026-04-16", "2:00 p.m. EST", "Envelope and Embodied Energy Subgroup", 1),
        ("2026-04-30", "2:00 p.m. EST", "Envelope and Embodied Energy Subgroup", 1),
        
        # CO HVACR — "Thursday pm (3/5, 3/19)"
        #   tentative: 4/16, 4/30
        ("2026-03-05", "2:00 p.m. EST", "Commercial HVACR and Water Heating Subgroup", 0),
        ("2026-03-19", "2:00 p.m. EST", "Commercial HVACR and Water Heating Subgroup", 0),
        ("2026-04-16", "2:00 p.m. EST", "Commercial HVACR and Water Heating Subgroup", 1),
        ("2026-04-30", "2:00 p.m. EST", "Commercial HVACR and Water Heating Subgroup", 1),
        
        # CO Modeling — "Thursday pm (3/12, 3/26, 4/16)"
        #   tentative: 4/6, 4/20
        ("2026-03-12", "2:00 p.m. EST", "Commercial Modeling Subgroup", 0),
        ("2026-03-26", "2:00 p.m. EST", "Commercial Modeling Subgroup", 0),
        ("2026-04-16", "2:00 p.m. EST", "Commercial Modeling Subgroup", 0),
        ("2026-04-06", "2:00 p.m. EST", "Commercial Modeling Subgroup", 1),
        ("2026-04-20", "2:00 p.m. EST", "Commercial Modeling Subgroup", 1),
        
        # ── RESIDENTIAL TRACK (for reference / future DB) ──
        
        # RE Consensus — Various
        # "Monday pm (2/24 - See additional Meeting information) (3/2)"
        # "Residential Consensus" also appears at: Thursday am (3/5, 3/19), Friday pm (4/10, 4/24)
        # Tentative: 4/23, 4/30
        ("2026-02-24", "2:00 p.m. EST", "Residential Consensus Committee", 0),
        ("2026-03-02", "2:00 p.m. EST", "Residential Consensus Committee", 0),
        ("2026-03-05", "11:00 a.m. EST", "Residential Consensus Committee", 0),
        ("2026-03-19", "11:00 a.m. EST", "Residential Consensus Committee", 0),
        ("2026-04-10", "2:00 p.m. EST", "Residential Consensus Committee", 0),
        ("2026-04-24", "2:00 p.m. EST", "Residential Consensus Committee", 0),
        ("2026-04-23", "2:00 p.m. EST", "Residential Consensus Committee", 1),
        ("2026-04-30", "2:00 p.m. EST", "Residential Consensus Committee", 1),
        
        # RE HVAC — "Monday pm (2/23, 3/23)" + added 2/23 meeting 11 a.m. to 2 p.m.
        ("2026-02-23", "11:00 a.m. to 2:00 p.m. EST", "Residential HVAC Subgroup", 0),
        ("2026-03-23", "2:00 p.m. EST", "Residential HVAC Subgroup", 0),
        ("2026-04-20", "2:00 p.m. EST", "Residential HVAC Subgroup", 1),
        
        # RE Envelope — "Monday am (2/23, 4/6)" + added 2/18 10 a.m. meeting
        ("2026-02-18", "10:00 a.m. EST", "Residential Envelope Subgroup", 0),
        ("2026-02-23", "11:00 a.m. EST", "Residential Envelope Subgroup", 0),
        ("2026-04-06", "11:00 a.m. EST", "Residential Envelope Subgroup", 0),
        ("2026-04-29", "11:00 a.m. EST", "Residential Envelope Subgroup", 1),
        
        # RE Existing Building — "Thursday pm (3/12, 3/26, 4/16 - 2:00 p.m. to 5:00 p.m. EST)"
        #   tentative: 4/28
        ("2026-03-12", "2:00 p.m. to 5:00 p.m. EST", "Residential Existing Building Subgroup", 0),
        ("2026-03-26", "2:00 p.m. to 5:00 p.m. EST", "Residential Existing Building Subgroup", 0),
        ("2026-04-16", "2:00 p.m. to 5:00 p.m. EST", "Residential Existing Building Subgroup", 0),
        ("2026-04-28", "2:00 p.m. to 5:00 p.m. EST", "Residential Existing Building Subgroup", 1),
        
        # RE EPLR — "Friday pm (4/3, 4/17)" + added 2/26 at 2:00 p.m.
        #   tentative: 4/13, 4/27
        ("2026-02-26", "2:00 p.m. EST", "Residential EPLR Subgroup", 0),
        ("2026-04-03", "2:00 p.m. EST", "Residential EPLR Subgroup", 0),
        ("2026-04-17", "2:00 p.m. EST", "Residential EPLR Subgroup", 0),
        ("2026-04-13", "2:00 p.m. EST", "Residential EPLR Subgroup", 1),
        ("2026-04-27", "2:00 p.m. EST", "Residential EPLR Subgroup", 1),
        
        # RE Modeling — "Tuesday am (3/24)" + tentative 4/14, 4/28
        ("2026-03-24", "11:00 a.m. EST", "Residential Modeling Subgroup", 0),
        ("2026-04-14", "11:00 a.m. EST", "Residential Modeling Subgroup", 1),
        ("2026-04-28", "11:00 a.m. EST", "Residential Modeling Subgroup", 1),
        
        # RE Consistency and Administration — "Tuesday pm (3/3)" + "Wednesday pm (2/25, 3/4, 3/11, 3/25)"
        # OCR: "Wednesday am (2/25 from 11 a.m. to 2 p.m. EST) (3/3, 3/12, 3/26, 4/16)"
        # This seems to be the RE Admin subgroup
        ("2026-03-03", "2:00 p.m. EST", "Residential Administration Subgroup", 0),
        ("2026-04-29", "2:00 p.m. EST", "Residential Administration Subgroup", 1),
    ]
    
    inserted = 0
    for meeting_date, meeting_time, body, tentative in schedule:
        # Determine status: if meeting_date is in the past relative to source date (2026-02-21), 
        # mark as SCHEDULED (we don't know if they happened yet)
        status = 'SCHEDULED'
        phase = 'PUBLIC_COMMENT'  # All these are in the PC phase
        
        conn.execute("""
            INSERT INTO meetings (meeting_date, meeting_time, body, phase, status, tentative, source)
            VALUES (?, ?, ?, ?, ?, ?, '260221_schedule_pdf')
            ON CONFLICT(meeting_date, body) DO UPDATE SET 
                meeting_time = COALESCE(excluded.meeting_time, meetings.meeting_time),
                tentative = excluded.tentative,
                source = CASE WHEN meetings.source = 'JSON_MINUTES' THEN meetings.source ELSE excluded.source END
        """, (meeting_date, meeting_time, body, phase, status, tentative))
        inserted += 1
    
    conn.commit()
    
    # Count what we have
    c = conn.cursor()
    total = c.execute("SELECT COUNT(*) FROM meetings").fetchone()[0]
    commercial = c.execute("SELECT COUNT(*) FROM meetings WHERE body LIKE 'Commercial%'").fetchone()[0]
    residential = c.execute("SELECT COUNT(*) FROM meetings WHERE body LIKE 'Residential%'").fetchone()[0]
    completed = c.execute("SELECT COUNT(*) FROM meetings WHERE status = 'COMPLETED'").fetchone()[0]
    scheduled = c.execute("SELECT COUNT(*) FROM meetings WHERE status = 'SCHEDULED'").fetchone()[0]
    tentative_ct = c.execute("SELECT COUNT(*) FROM meetings WHERE tentative = 1").fetchone()[0]
    
    print(f"    → {total} meetings total ({commercial} commercial, {residential} residential)")
    print(f"    → {completed} completed, {scheduled} scheduled ({tentative_ct} tentative)")
    
    # Print commercial consensus schedule
    print("\n    Commercial Consensus Committee schedule:")
    for row in c.execute("""
        SELECT meeting_date, meeting_time, status, tentative 
        FROM meetings WHERE body = 'Commercial Consensus Committee' 
        ORDER BY meeting_date
    """):
        tent = " (TENTATIVE)" if row[3] else ""
        print(f"      {row[0]}  {row[1] or ''}  [{row[2]}]{tent}")

# ─── CREATE VIEWS ────────────────────────────────────────────────────────────

def create_views(conn):
    print("\n[7] Creating views...")
    conn.executescript("""
    
    -- Full status view: every proposal with its final disposition
    CREATE VIEW v_current_status AS
    SELECT 
        p.canonical_id,
        p.cycle,
        p.phase,
        p.part,
        p.code_section,
        p.proponent,
        p.initial_subgroup,
        p.current_subgroup,
        p.cdpaccess_id,
        p.cdpaccess_url,
        p.withdrawn,
        sa.subgroup AS sg_subgroup,
        sa.recommendation AS sg_recommendation,
        sa.action_date AS sg_action_date,
        sa.vote_for AS sg_vote_for,
        sa.vote_against AS sg_vote_against,
        sa.vote_not_voting AS sg_vote_not_voting,
        sa.reason AS sg_reason,
        sa.modification_text AS sg_modification,
        ca.recommendation AS consensus_recommendation,
        ca.action_date AS consensus_date,
        ca.vote_for AS cons_vote_for,
        ca.vote_against AS cons_vote_against,
        ca.vote_not_voting AS cons_vote_not_voting,
        ca.reason AS cons_reason,
        ca.modification_text AS cons_modification,
        ca.moved_by,
        ca.seconded_by,
        CASE 
            WHEN p.withdrawn = 1 THEN 'WITHDRAWN'
            WHEN ca.recommendation IS NOT NULL THEN 'DECIDED'
            WHEN sa.recommendation IS NOT NULL THEN 'PENDING_CONSENSUS'
            ELSE 'NO_ACTION'
        END AS status
    FROM proposals p
    LEFT JOIN subgroup_actions sa ON p.proposal_uid = sa.proposal_uid
    LEFT JOIN consensus_actions ca ON p.proposal_uid = ca.proposal_uid AND ca.is_final = 1
    ORDER BY p.canonical_id;

    -- Ready for consensus: Public Comment proposals with subgroup action but no final consensus
    CREATE VIEW v_ready_for_consensus AS
    SELECT 
        p.canonical_id,
        p.code_section,
        p.proponent,
        p.current_subgroup,
        sa.subgroup,
        sa.recommendation AS sg_recommendation,
        sa.action_date AS sg_action_date,
        sa.vote_for AS sg_vote_for,
        sa.vote_against AS sg_vote_against,
        sa.vote_not_voting AS sg_vote_not_voting,
        sa.reason AS sg_reason,
        sa.modification_text AS sg_modification
    FROM proposals p
    INNER JOIN subgroup_actions sa ON p.proposal_uid = sa.proposal_uid
    LEFT JOIN consensus_actions ca ON p.proposal_uid = ca.proposal_uid AND ca.is_final = 1
    WHERE p.phase = 'PUBLIC_COMMENT'
      AND p.withdrawn = 0
      AND ca.id IS NULL
    ORDER BY sa.subgroup, p.canonical_id;

    -- Data quality review
    CREATE VIEW v_data_quality_review AS
    SELECT 
        dqf.flag_type,
        dqf.canonical_id,
        dqf.table_name,
        dqf.raw_value,
        dqf.resolved_value,
        dqf.needs_review,
        dqf.created_at
    FROM data_quality_flags dqf
    ORDER BY dqf.needs_review DESC, dqf.flag_type, dqf.canonical_id;

    -- Multi-action proposals (procedural sequences)
    CREATE VIEW v_multi_action_proposals AS
    SELECT 
        p.canonical_id,
        ca.sequence,
        ca.action_date,
        ca.recommendation,
        ca.recommendation_raw,
        ca.vote_for,
        ca.vote_against,
        ca.moved_by,
        ca.is_final,
        ca.source
    FROM consensus_actions ca
    JOIN proposals p ON ca.proposal_uid = p.proposal_uid
    WHERE p.proposal_uid IN (
        SELECT proposal_uid FROM consensus_actions GROUP BY proposal_uid HAVING COUNT(*) > 1
    )
    ORDER BY p.canonical_id, ca.sequence;
    """)
    conn.commit()
    print("    → 4 views created")

# ─── VERIFICATION ────────────────────────────────────────────────────────────

def verify(conn):
    print("\n" + "="*70)
    print("VERIFICATION REPORT")
    print("="*70)
    
    c = conn.cursor()
    
    # Basic counts
    total = c.execute("SELECT COUNT(*) FROM proposals").fetchone()[0]
    ce = c.execute("SELECT COUNT(*) FROM proposals WHERE phase = 'CE'").fetchone()[0]
    pc = c.execute("SELECT COUNT(*) FROM proposals WHERE phase = 'PUBLIC_COMMENT'").fetchone()[0]
    print(f"\nProposals: {total} total ({ce} CE, {pc} PUBLIC_COMMENT)")
    
    withdrawn = c.execute("SELECT COUNT(*) FROM proposals WHERE withdrawn = 1").fetchone()[0]
    print(f"Withdrawn: {withdrawn}")
    
    with_cdp = c.execute("SELECT COUNT(*) FROM proposals WHERE cdpaccess_id IS NOT NULL").fetchone()[0]
    print(f"With cdpACCESS ID: {with_cdp}")
    
    # Subgroup actions
    sg_count = c.execute("SELECT COUNT(*) FROM subgroup_actions").fetchone()[0]
    print(f"\nSubgroup actions: {sg_count}")
    sg_by_group = c.execute("SELECT subgroup, COUNT(*) FROM subgroup_actions GROUP BY subgroup ORDER BY subgroup").fetchall()
    for sg, cnt in sg_by_group:
        print(f"  {sg}: {cnt}")
    
    # Consensus actions
    ca_count = c.execute("SELECT COUNT(*) FROM consensus_actions").fetchone()[0]
    ca_final = c.execute("SELECT COUNT(*) FROM consensus_actions WHERE is_final = 1").fetchone()[0]
    ca_json = c.execute("SELECT COUNT(*) FROM consensus_actions WHERE source = 'JSON_MINUTES'").fetchone()[0]
    ca_excel = c.execute("SELECT COUNT(*) FROM consensus_actions WHERE source = 'EXCEL_TRACKING'").fetchone()[0]
    print(f"\nConsensus actions: {ca_count} total ({ca_final} final)")
    print(f"  From JSON minutes: {ca_json}")
    print(f"  From Excel tracking: {ca_excel}")
    
    # Meetings
    meetings = c.execute("SELECT meeting_date, body, status, action_count, tentative FROM meetings ORDER BY meeting_date").fetchall()
    completed_mtgs = [m for m in meetings if m[2] == 'COMPLETED']
    scheduled_mtgs = [m for m in meetings if m[2] == 'SCHEDULED' and not m[4]]
    tentative_mtgs = [m for m in meetings if m[4]]
    print(f"\nMeetings: {len(meetings)} total ({len(completed_mtgs)} completed, {len(scheduled_mtgs)} scheduled, {len(tentative_mtgs)} tentative)")
    print("  Completed (with actions):")
    for dt, body, status, cnt, tent in completed_mtgs:
        print(f"    {dt} | {body}: {cnt} actions")
    
    # Next commercial consensus meeting
    next_cc = c.execute("""
        SELECT meeting_date, meeting_time FROM meetings 
        WHERE body = 'Commercial Consensus Committee' AND meeting_date >= date('now')
        ORDER BY meeting_date LIMIT 1
    """).fetchone()
    if next_cc:
        print(f"\n  ► NEXT Commercial Consensus: {next_cc[0]} at {next_cc[1]}")
    
    # Status breakdown
    print("\nStatus breakdown:")
    statuses = c.execute("""
        SELECT status, COUNT(*) FROM v_current_status GROUP BY status ORDER BY status
    """).fetchall()
    for status, cnt in statuses:
        print(f"  {status}: {cnt}")
    
    # Ready for consensus
    ready = c.execute("SELECT COUNT(*) FROM v_ready_for_consensus").fetchone()[0]
    print(f"\nReady for consensus: {ready}")
    ready_by_sg = c.execute("SELECT subgroup, COUNT(*) FROM v_ready_for_consensus GROUP BY subgroup").fetchall()
    for sg, cnt in ready_by_sg:
        print(f"  {sg}: {cnt}")
    
    # Multi-action proposals
    multi = c.execute("""
        SELECT COUNT(DISTINCT proposal_uid) FROM consensus_actions 
        GROUP BY proposal_uid HAVING COUNT(*) > 1
    """).fetchall()
    print(f"\nMulti-action proposals: {len(multi)}")
    
    # Movements
    mvmt = c.execute("SELECT COUNT(*) FROM subgroup_movements").fetchone()[0]
    print(f"Subgroup movements: {mvmt}")
    
    # Data quality
    flags = c.execute("SELECT flag_type, COUNT(*) FROM data_quality_flags GROUP BY flag_type ORDER BY COUNT(*) DESC").fetchall()
    print(f"\nData quality flags: {sum(cnt for _, cnt in flags)} total")
    for ft, cnt in flags:
        print(f"  {ft}: {cnt}")
    
    # Integrity checks
    print("\n--- INTEGRITY CHECKS ---")
    
    # Orphan subgroup actions
    orphan_sa = c.execute("""
        SELECT COUNT(*) FROM subgroup_actions sa 
        LEFT JOIN proposals p ON sa.proposal_uid = p.proposal_uid 
        WHERE p.proposal_uid IS NULL
    """).fetchone()[0]
    print(f"Orphan subgroup actions: {orphan_sa} {'✓' if orphan_sa == 0 else '✗ PROBLEM'}")
    
    # Orphan consensus actions
    orphan_ca = c.execute("""
        SELECT COUNT(*) FROM consensus_actions ca 
        LEFT JOIN proposals p ON ca.proposal_uid = p.proposal_uid 
        WHERE p.proposal_uid IS NULL
    """).fetchone()[0]
    print(f"Orphan consensus actions: {orphan_ca} {'✓' if orphan_ca == 0 else '✗ PROBLEM'}")
    
    # CE proposals in ready-for-consensus (should be 0)
    ce_in_ready = c.execute("SELECT COUNT(*) FROM v_ready_for_consensus WHERE canonical_id LIKE 'CE%' AND canonical_id NOT LIKE 'CEPC%' AND canonical_id NOT LIKE 'CECP%' AND canonical_id NOT LIKE 'CECC%'").fetchone()[0]
    print(f"CE proposals in ready queue: {ce_in_ready} {'✓' if ce_in_ready == 0 else '✗ PROBLEM'}")
    
    # Duplicate canonical IDs
    dupes = c.execute("SELECT canonical_id, COUNT(*) FROM proposals GROUP BY canonical_id HAVING COUNT(*) > 1").fetchall()
    print(f"Duplicate canonical IDs: {len(dupes)} {'✓' if len(dupes) == 0 else '✗ PROBLEM'}")
    for cid, cnt in dupes:
        print(f"  {cid}: {cnt}")

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    print("="*70)
    print("IECC COMMERCIAL DATABASE BUILD")
    print(f"Started: {datetime.now().isoformat()}")
    print("="*70)
    
    # Remove old DB
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    
    create_db(conn)
    ingest_cdpaccess(conn)
    ingest_ce_tracking(conn)
    ingest_pc_tracking(conn)
    ingest_sort_pass(conn)
    ingest_subgroup_jsons(conn)
    ingest_consensus_json(conn)
    ingest_schedule(conn)
    create_views(conn)
    verify(conn)
    
    conn.close()
    
    size_mb = os.path.getsize(DB_PATH) / (1024*1024)
    print(f"\n{'='*70}")
    print(f"Database written: {DB_PATH} ({size_mb:.2f} MB)")
    print(f"Completed: {datetime.now().isoformat()}")
    print(f"{'='*70}")

if __name__ == "__main__":
    main()
