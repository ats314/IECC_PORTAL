#!/usr/bin/env python3
"""Generate IECC_2027_Combined_Disposition.xlsx from unified iecc.db.

Single master report. Sheets:
  1. Dashboard — status summary, coverage, pending breakdown, decisions, ready/awaiting lists
  2. Commercial Proposals — full disposition with status
  3. Residential Proposals — full disposition with status
  4. Commercial CA — every consensus action record (with sequence, mover, source)
  5. Residential CA — every consensus action record
  6. Commercial SA — every subgroup action record (with reason, source)
  7. Residential SA — every subgroup action record
  8. Crossovers — proposals that cross track boundaries
  9. Data Quality — all DQ flags with resolved values
  10. Errata — errata records
  11. Meetings — all meeting records
"""
import sqlite3
import os
from pathlib import Path
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
DB_PATH = str(BASE / "iecc.db")
OUTPUT = str(BASE / "IECC_2027_Combined_Disposition.xlsx")

# --- Colors ---
DARK_BLUE = "2F5496"
MED_BLUE = "4472C4"
NAVY = "1F4E79"
DARK_GREEN = "2E7D32"
ACCENT_ORANGE = "E65100"
ACCENT_RED = "B71C1C"
ACCENT_PURPLE = "6A1B9A"
DARK_GRAY = "333333"
LIGHT_GRAY = "F5F5F5"
WHITE = "FFFFFF"

# --- Fonts ---
TITLE_FONT = Font(name="Arial", bold=True, size=14, color=WHITE)
SUBTITLE_FONT = Font(name="Arial", bold=True, size=9, italic=True, color=WHITE)
SECTION_FONT = Font(name="Arial", bold=True, size=11, color=WHITE)
HEADER_FONT = Font(name="Arial", bold=True, size=10, color=WHITE)
METRIC_FONT = Font(name="Arial", size=10)
METRIC_BOLD = Font(name="Arial", bold=True, size=10)
VALUE_FONT = Font(name="Arial", size=10)
VALUE_BOLD = Font(name="Arial", bold=True, size=11)
DATA_FONT = Font(name="Arial", size=10)
NUM_FONT = Font(name="Arial", size=11)

# --- Fills ---
TITLE_FILL = PatternFill("solid", fgColor=DARK_BLUE)
SECTION_FILL = PatternFill("solid", fgColor=MED_BLUE)
HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
HEADER_FILL_COM = PatternFill("solid", fgColor=NAVY)
HEADER_FILL_RES = PatternFill("solid", fgColor=DARK_GREEN)
HEADER_FILL_DQ = PatternFill("solid", fgColor=ACCENT_RED)
HEADER_FILL_ERRATA = PatternFill("solid", fgColor=ACCENT_PURPLE)
HEADER_FILL_MTG = PatternFill("solid", fgColor=ACCENT_ORANGE)
HEADER_FILL_DASH = PatternFill("solid", fgColor=DARK_GRAY)
PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")
WITHDRAWN_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL = PatternFill("solid", fgColor="FCE4EC")
ORANGE_FILL = PatternFill("solid", fgColor="FBE5D6")
YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")
GRAY_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

PCT_FMT = '0%'
CENTER = Alignment(horizontal="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP = Alignment(vertical="top", wrap_text=True)


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def query(sql, params=()):
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def scalar(sql, params=()):
    conn = sqlite3.connect(DB_PATH)
    val = conn.execute(sql, params).fetchone()[0]
    conn.close()
    return val


def wcell(ws, row, col, value, font=None, fill=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    c.border = THIN_BORDER
    return c


def style_header_row(ws, row, ncols, fill=None):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = fill or HEADER_FILL
        c.alignment = CENTER_WRAP
        c.border = THIN_BORDER


def style_section_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = SECTION_FONT
        c.fill = SECTION_FILL
        c.alignment = Alignment(horizontal="left")
        c.border = THIN_BORDER


def merge_section(ws, row, col_end, value):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=value)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, col_end + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER


def style_data_row(ws, row, ncols, status=None, alt=False):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = DATA_FONT
        c.border = THIN_BORDER
        c.alignment = WRAP
        if alt:
            c.fill = GRAY_FILL
        if status == "Pending":
            c.fill = PENDING_FILL
        elif status == "Withdrawn":
            c.fill = WITHDRAWN_FILL


def auto_width(ws, min_w=10, max_w=45):
    for col_cells in ws.columns:
        lengths = []
        for cell in col_cells:
            if cell.value:
                lengths.append(min(len(str(cell.value)), max_w))
        if lengths:
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(min_w, max(lengths) + 2)


def format_vote(vf, va, vnv=None):
    parts = []
    for v in [vf, va, vnv]:
        if v is not None:
            parts.append(str(int(v)))
        else:
            parts.append('-')
    return '-'.join(parts) if any(p != '-' for p in parts) else ''


# ==================== DASHBOARD ====================
def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = DARK_BLUE
    NCOLS = 5
    conn = get_conn()

    stats = {}
    for track in ["commercial", "residential"]:
        d = {}
        d["total"] = conn.execute("SELECT COUNT(*) FROM proposals WHERE track=?", (track,)).fetchone()[0]
        for s in ["Decided", "Pending", "Withdrawn", "Phase Closed"]:
            d[s] = conn.execute("SELECT COUNT(*) FROM v_current_status WHERE track=? AND status=?", (track, s)).fetchone()[0]
        d["ca_total"] = conn.execute("SELECT COUNT(*) FROM consensus_actions ca JOIN proposals p ON ca.proposal_uid=p.proposal_uid WHERE p.track=?", (track,)).fetchone()[0]
        d["ca_final"] = conn.execute("SELECT COUNT(*) FROM consensus_actions ca JOIN proposals p ON ca.proposal_uid=p.proposal_uid WHERE p.track=? AND ca.is_final=1", (track,)).fetchone()[0]
        d["ca_votes"] = conn.execute("SELECT COUNT(*) FROM consensus_actions ca JOIN proposals p ON ca.proposal_uid=p.proposal_uid WHERE p.track=? AND ca.is_final=1 AND ca.vote_for IS NOT NULL", (track,)).fetchone()[0]
        d["ca_reasons"] = conn.execute("SELECT COUNT(*) FROM consensus_actions ca JOIN proposals p ON ca.proposal_uid=p.proposal_uid WHERE p.track=? AND ca.is_final=1 AND ca.reason IS NOT NULL AND ca.reason != ''", (track,)).fetchone()[0]
        d["ca_dates"] = conn.execute("SELECT COUNT(*) FROM consensus_actions ca JOIN proposals p ON ca.proposal_uid=p.proposal_uid WHERE p.track=? AND ca.is_final=1 AND ca.action_date IS NOT NULL", (track,)).fetchone()[0]
        d["sa"] = conn.execute("SELECT COUNT(*) FROM subgroup_actions sa JOIN proposals p ON sa.proposal_uid=p.proposal_uid WHERE p.track=?", (track,)).fetchone()[0]
        d["sa_votes"] = conn.execute("SELECT COUNT(*) FROM subgroup_actions sa JOIN proposals p ON sa.proposal_uid=p.proposal_uid WHERE p.track=? AND sa.vote_for IS NOT NULL", (track,)).fetchone()[0]
        d["sa_reasons"] = conn.execute("SELECT COUNT(*) FROM subgroup_actions sa JOIN proposals p ON sa.proposal_uid=p.proposal_uid WHERE p.track=? AND sa.reason IS NOT NULL AND sa.reason != ''", (track,)).fetchone()[0]
        d["dq_open"] = conn.execute("SELECT COUNT(*) FROM data_quality_flags WHERE track=? AND needs_review=1", (track,)).fetchone()[0]
        d["dq_total"] = conn.execute("SELECT COUNT(*) FROM data_quality_flags WHERE track=?", (track,)).fetchone()[0]
        stats[track] = d

    cm = stats["commercial"]
    rs = stats["residential"]
    row = 1

    # === TITLE ===
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    wcell(ws, row, 1, "IECC 2027 Secretariat — Master Report", TITLE_FONT, TITLE_FILL)
    for col in range(2, NCOLS + 1):
        ws.cell(row=row, column=col).fill = TITLE_FILL
    ws.row_dimensions[row].height = 30
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    wcell(ws, row, 1, f"Generated {date.today().isoformat()}  |  Source: iecc.db (unified)  |  Deadline: April 30, 2026", SUBTITLE_FONT, TITLE_FILL)
    for col in range(2, NCOLS + 1):
        ws.cell(row=row, column=col).fill = TITLE_FILL
    row += 2

    # === SECTION 1: OVERALL STATUS ===
    merge_section(ws, row, NCOLS, "Overall Status")
    row += 1
    for i, h in enumerate(["", "Commercial", "Residential", "Combined", ""], 1):
        wcell(ws, row, i, h, HEADER_FONT, HEADER_FILL, CENTER)
    row += 1

    for name, key in [("Total Proposals", "total"), ("Decided", "Decided"), ("Pending", "Pending"),
                       ("Withdrawn", "Withdrawn"), ("Phase Closed", "Phase Closed")]:
        cv, rv = cm[key], rs[key]
        wcell(ws, row, 1, name, METRIC_BOLD if key == "total" else METRIC_FONT)
        wcell(ws, row, 2, cv, VALUE_FONT, align=CENTER)
        wcell(ws, row, 3, rv, VALUE_FONT, align=CENTER)
        wcell(ws, row, 4, cv + rv, VALUE_BOLD, align=CENTER)
        if key == "Decided":
            pct = (cv + rv) / (cm["total"] + rs["total"])
            wcell(ws, row, 5, pct, VALUE_FONT, GREEN_FILL if pct >= 0.7 else ORANGE_FILL, CENTER, PCT_FMT)
        elif key == "Pending":
            pct = (cv + rv) / (cm["total"] + rs["total"])
            wcell(ws, row, 5, pct, VALUE_FONT, ORANGE_FILL if pct > 0.05 else GREEN_FILL, CENTER, PCT_FMT)
        else:
            wcell(ws, row, 5, "")
        row += 1
    row += 1

    # === SECTION 2: DATA COVERAGE ===
    merge_section(ws, row, NCOLS, "Data Coverage")
    row += 1
    for i, h in enumerate(["", "Commercial", "Residential", "Combined", ""], 1):
        wcell(ws, row, i, h, HEADER_FONT, HEADER_FILL, CENTER)
    row += 1

    cov_rows = [
        ("Consensus Actions (all)", "ca_total"),
        ("  Final CA", "ca_final"),
        ("  - with vote counts", "ca_votes"),
        ("  - with reasons", "ca_reasons"),
        ("  - with dates", "ca_dates"),
        ("Subgroup Actions", "sa"),
        ("  - with vote counts", "sa_votes"),
        ("  - with reasons", "sa_reasons"),
    ]
    for name, key in cov_rows:
        cv, rv = cm[key], rs[key]
        wcell(ws, row, 1, name, METRIC_FONT)
        wcell(ws, row, 2, cv, VALUE_FONT, align=CENTER)
        wcell(ws, row, 3, rv, VALUE_FONT, align=CENTER)
        wcell(ws, row, 4, cv + rv, VALUE_BOLD, align=CENTER)
        # Show coverage % for sub-items
        if key in ("ca_votes", "ca_reasons", "ca_dates"):
            base_c, base_r = cm["ca_final"], rs["ca_final"]
            if base_c + base_r > 0:
                pct = (cv + rv) / (base_c + base_r)
                wcell(ws, row, 5, pct, VALUE_FONT, GREEN_FILL if pct >= 0.9 else ORANGE_FILL, CENTER, PCT_FMT)
        elif key in ("sa_votes", "sa_reasons"):
            base_c, base_r = cm["sa"], rs["sa"]
            if base_c + base_r > 0:
                pct = (cv + rv) / (base_c + base_r)
                wcell(ws, row, 5, pct, VALUE_FONT, GREEN_FILL if pct >= 0.9 else ORANGE_FILL, CENTER, PCT_FMT)
        else:
            wcell(ws, row, 5, "")
        row += 1

    # DQ summary row
    wcell(ws, row, 1, "DQ Flags (open / total)", METRIC_FONT)
    wcell(ws, row, 2, f"{cm['dq_open']} / {cm['dq_total']}", VALUE_FONT, align=CENTER)
    wcell(ws, row, 3, f"{rs['dq_open']} / {rs['dq_total']}", VALUE_FONT, align=CENTER)
    wcell(ws, row, 4, f"{cm['dq_open']+rs['dq_open']} / {cm['dq_total']+rs['dq_total']}", VALUE_BOLD, align=CENTER)
    wcell(ws, row, 5, "")
    row += 2

    # === SECTION 3: PENDING BY SUBGROUP ===
    merge_section(ws, row, NCOLS, "Pending Proposals by Subgroup")
    row += 1
    for i, h in enumerate(["Track", "Subgroup", "Pending", "Ready for CC", "Awaiting SG"], 1):
        wcell(ws, row, i, h, HEADER_FONT, HEADER_FILL, CENTER)
    row += 1

    for track in ["commercial", "residential"]:
        sgs = conn.execute("""
            SELECT current_subgroup,
                   COUNT(*) as total,
                   SUM(CASE WHEN sg_recommendation IS NOT NULL THEN 1 ELSE 0 END) as ready,
                   SUM(CASE WHEN sg_recommendation IS NULL THEN 1 ELSE 0 END) as awaiting
            FROM v_current_status WHERE track=? AND status='Pending'
            GROUP BY current_subgroup ORDER BY COUNT(*) DESC
        """, (track,)).fetchall()
        for sg in sgs:
            wcell(ws, row, 1, track.capitalize(), METRIC_FONT, align=CENTER)
            wcell(ws, row, 2, sg["current_subgroup"], METRIC_FONT)
            wcell(ws, row, 3, sg["total"], VALUE_FONT, align=CENTER)
            wcell(ws, row, 4, sg["ready"], VALUE_FONT, GREEN_FILL, CENTER)
            aw = sg["awaiting"]
            wcell(ws, row, 5, aw, VALUE_FONT, ORANGE_FILL if aw > 0 else GREEN_FILL, CENTER)
            row += 1
    row += 1

    # === SECTION 4: DECISION BREAKDOWN ===
    merge_section(ws, row, NCOLS, "Decision Breakdown (Decided Proposals)")
    row += 1
    for i, h in enumerate(["Decision", "Commercial", "Residential", "Combined", "% of Decided"], 1):
        wcell(ws, row, i, h, HEADER_FONT, HEADER_FILL, CENTER)
    row += 1

    decisions = conn.execute("""
        SELECT ca_recommendation,
               SUM(CASE WHEN track='commercial' THEN 1 ELSE 0 END) as comm,
               SUM(CASE WHEN track='residential' THEN 1 ELSE 0 END) as res,
               COUNT(*) as total
        FROM v_current_status WHERE status='Decided' AND ca_recommendation IS NOT NULL
        GROUP BY ca_recommendation ORDER BY COUNT(*) DESC
    """).fetchall()
    total_decided = cm["Decided"] + rs["Decided"]
    for d in decisions:
        wcell(ws, row, 1, d["ca_recommendation"], METRIC_FONT)
        wcell(ws, row, 2, d["comm"], VALUE_FONT, align=CENTER)
        wcell(ws, row, 3, d["res"], VALUE_FONT, align=CENTER)
        wcell(ws, row, 4, d["total"], VALUE_BOLD, align=CENTER)
        pct = d["total"] / total_decided if total_decided else 0
        wcell(ws, row, 5, pct, VALUE_FONT, align=CENTER, fmt=PCT_FMT)
        row += 1
    row += 1

    # === SECTION 5: READY FOR CONSENSUS ===
    merge_section(ws, row, NCOLS, "Ready for Consensus Committee (have SG rec, no CA yet)")
    row += 1
    for i, h in enumerate(["Track", "Proposal", "Subgroup", "SG Rec", "SG Vote"], 1):
        wcell(ws, row, i, h, HEADER_FONT, HEADER_FILL, CENTER)
    row += 1

    ready = conn.execute("""
        SELECT track, canonical_id, current_subgroup, sg_recommendation, sg_vote_for, sg_vote_against
        FROM v_current_status WHERE status='Pending' AND sg_recommendation IS NOT NULL
        ORDER BY track, current_subgroup, canonical_id
    """).fetchall()
    for p in ready:
        vote = f"{p['sg_vote_for']}-{p['sg_vote_against']}" if p['sg_vote_for'] is not None else ""
        wcell(ws, row, 1, p["track"].capitalize(), METRIC_FONT, align=CENTER)
        wcell(ws, row, 2, p["canonical_id"], METRIC_BOLD)
        wcell(ws, row, 3, p["current_subgroup"], METRIC_FONT)
        wcell(ws, row, 4, p["sg_recommendation"], METRIC_FONT, GREEN_FILL)
        wcell(ws, row, 5, vote, METRIC_FONT, align=CENTER)
        row += 1
    row += 1

    # === SECTION 6: AWAITING SG ACTION ===
    merge_section(ws, row, NCOLS, "Awaiting Subgroup Action (no SG recommendation yet)")
    row += 1
    for i, h in enumerate(["Track", "Proposal", "Subgroup", "Proponent", ""], 1):
        wcell(ws, row, i, h, HEADER_FONT, HEADER_FILL, CENTER)
    row += 1

    awaiting = conn.execute("""
        SELECT track, canonical_id, current_subgroup, proponent
        FROM v_current_status WHERE status='Pending' AND sg_recommendation IS NULL
        ORDER BY track, current_subgroup, canonical_id
    """).fetchall()
    for p in awaiting:
        wcell(ws, row, 1, p["track"].capitalize(), METRIC_FONT, align=CENTER)
        wcell(ws, row, 2, p["canonical_id"], METRIC_BOLD, ORANGE_FILL)
        wcell(ws, row, 3, p["current_subgroup"], METRIC_FONT)
        wcell(ws, row, 4, p["proponent"] or "", METRIC_FONT)
        wcell(ws, row, 5, "")
        row += 1
    row += 1

    # === SECTION 7: KEY DATES ===
    merge_section(ws, row, NCOLS, "Key Dates")
    row += 1
    dates = [
        ("Committee Action Deadline", "April 30, 2026"),
        ("CAR Issued", "May 7, 2026"),
        ("Commenter Objection Deadline", "June 7, 2026"),
        ("Final Draft 2027 IECC", "December 1, 2026"),
    ]
    for label, val in dates:
        wcell(ws, row, 1, label, METRIC_FONT)
        wcell(ws, row, 2, val, VALUE_FONT)
        for col in range(3, NCOLS + 1):
            wcell(ws, row, col, "")
        row += 1

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 16
    ws.sheet_view.showGridLines = False
    conn.close()


# ==================== PROPOSAL SHEETS ====================
def build_proposal_sheet(wb, track, sheet_name, fill):
    conn = get_conn()
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = NAVY if track == "commercial" else DARK_GREEN

    headers = ["Proposal", "Phase", "Proponent", "Subgroup", "SG Rec", "SG Vote",
               "SG Date", "CA Decision", "CA Vote", "CA Date", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), fill)
    ws.freeze_panes = "A2"

    rows = conn.execute("""
        SELECT canonical_id, phase, proponent, current_subgroup,
               sg_recommendation, sg_vote_for, sg_vote_against, sg_date,
               ca_recommendation, ca_vote_for, ca_vote_against, ca_date, status
        FROM v_current_status WHERE track=? ORDER BY canonical_id
    """, (track,)).fetchall()

    for r_idx, r in enumerate(rows, 2):
        sg_vote = format_vote(r['sg_vote_for'], r['sg_vote_against'])
        ca_vote = format_vote(r['ca_vote_for'], r['ca_vote_against'])
        vals = [r["canonical_id"], r["phase"], r["proponent"], r["current_subgroup"],
                r["sg_recommendation"], sg_vote, r["sg_date"],
                r["ca_recommendation"], ca_vote, r["ca_date"], r["status"]]
        for c_idx, val in enumerate(vals, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
        style_data_row(ws, r_idx, len(headers), status=r["status"])
        # Color status cell
        sc = ws.cell(row=r_idx, column=11)
        if r["status"] == "Decided":
            sc.fill = GREEN_FILL
        elif r["status"] == "Pending":
            sc.fill = YELLOW_FILL
        elif r["status"] == "Withdrawn":
            sc.fill = RED_FILL

    auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    conn.close()
    return len(rows)


# ==================== CA SHEETS ====================
def build_ca_sheet(wb, track, sheet_name, fill):
    conn = get_conn()
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = NAVY if track == "commercial" else DARK_GREEN

    headers = ["Proposal", "Seq", "Date", "Recommendation", "For", "Against", "NV",
               "Reason", "Moved By", "Seconded By", "Final", "Source"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), fill)
    ws.freeze_panes = "A2"

    rows = conn.execute("""
        SELECT p.canonical_id, ca.sequence, ca.action_date, ca.recommendation,
               ca.vote_for, ca.vote_against, ca.vote_not_voting,
               ca.reason, ca.moved_by, ca.seconded_by, ca.is_final, ca.source
        FROM consensus_actions ca JOIN proposals p ON ca.proposal_uid=p.proposal_uid
        WHERE p.track=? ORDER BY p.canonical_id, ca.sequence
    """, (track,)).fetchall()

    for i, r in enumerate(rows):
        rn = i + 2
        vals = [r["canonical_id"], r["sequence"], r["action_date"], r["recommendation"],
                r["vote_for"], r["vote_against"], r["vote_not_voting"],
                r["reason"], r["moved_by"], r["seconded_by"],
                "Yes" if r["is_final"] else "No", r["source"]]
        for c_idx, val in enumerate(vals, 1):
            ws.cell(row=rn, column=c_idx, value=val)
        style_data_row(ws, rn, len(headers), alt=(i % 2 == 1))

    auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    conn.close()
    return len(rows)


# ==================== SA SHEETS ====================
def build_sa_sheet(wb, track, sheet_name, fill):
    conn = get_conn()
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = NAVY if track == "commercial" else DARK_GREEN

    headers = ["Proposal", "Subgroup", "Date", "Recommendation", "For", "Against", "NV",
               "Reason", "Source"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), fill)
    ws.freeze_panes = "A2"

    rows = conn.execute("""
        SELECT p.canonical_id, sa.subgroup, sa.action_date, sa.recommendation,
               sa.vote_for, sa.vote_against, sa.vote_not_voting, sa.reason, sa.source_file
        FROM subgroup_actions sa JOIN proposals p ON sa.proposal_uid=p.proposal_uid
        WHERE p.track=? ORDER BY p.canonical_id
    """, (track,)).fetchall()

    for i, r in enumerate(rows):
        rn = i + 2
        vals = [r["canonical_id"], r["subgroup"], r["action_date"], r["recommendation"],
                r["vote_for"], r["vote_against"], r["vote_not_voting"],
                r["reason"], r["source_file"]]
        for c_idx, val in enumerate(vals, 1):
            ws.cell(row=rn, column=c_idx, value=val)
        style_data_row(ws, rn, len(headers), alt=(i % 2 == 1))

    auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    conn.close()
    return len(rows)


# ==================== CROSSOVERS ====================
def build_crossovers(wb):
    ws = wb.create_sheet("Crossovers")
    ws.sheet_properties.tabColor = ACCENT_PURPLE
    headers = ["Proposal", "Residential Status", "Res Vote", "Commercial Status", "Comm Vote"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers))

    conn = get_conn()
    crossovers = conn.execute("""
        SELECT p.canonical_id, ca.recommendation, ca.vote_for, ca.vote_against
        FROM proposals p
        LEFT JOIN consensus_actions ca ON p.proposal_uid=ca.proposal_uid AND ca.is_final=1
        WHERE p.track='residential' AND p.prefix IN ('CE','CEPC')
        ORDER BY p.canonical_id
    """).fetchall()

    for r_idx, row in enumerate(crossovers, 2):
        canon = row["canonical_id"]
        res_status = row["recommendation"] or "No action"
        res_vote = format_vote(row["vote_for"], row["vote_against"])
        comm_row = conn.execute("""
            SELECT ca.recommendation, ca.vote_for, ca.vote_against
            FROM proposals p LEFT JOIN consensus_actions ca ON p.proposal_uid=ca.proposal_uid AND ca.is_final=1
            WHERE p.track='commercial' AND p.canonical_id=?
        """, (canon,)).fetchone()
        comm_status = comm_row["recommendation"] if comm_row and comm_row["recommendation"] else "Not in commercial DB"
        comm_vote = format_vote(comm_row["vote_for"], comm_row["vote_against"]) if comm_row else ""
        vals = [canon, res_status, res_vote, comm_status, comm_vote]
        for c_idx, val in enumerate(vals, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
        style_data_row(ws, r_idx, len(headers))

    auto_width(ws)
    conn.close()
    return len(crossovers)


# ==================== DATA QUALITY ====================
def build_data_quality(wb):
    ws = wb.create_sheet("Data Quality")
    ws.sheet_properties.tabColor = ACCENT_RED
    headers = ["Track", "Proposal", "Table", "Flag Type", "Raw Value", "Resolved Value", "Needs Review"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), HEADER_FILL_DQ)
    ws.freeze_panes = "A2"

    conn = get_conn()
    rows = conn.execute("SELECT * FROM data_quality_flags ORDER BY track DESC, needs_review DESC, flag_type, canonical_id").fetchall()
    for i, r in enumerate(rows):
        rn = i + 2
        vals = [r["track"].capitalize(), r["canonical_id"], r["table_name"], r["flag_type"],
                r["raw_value"], r["resolved_value"], "Yes" if r["needs_review"] else "No"]
        for c_idx, val in enumerate(vals, 1):
            ws.cell(row=rn, column=c_idx, value=val)
        style_data_row(ws, rn, len(headers), alt=(i % 2 == 1))
        if r["needs_review"]:
            ws.cell(row=rn, column=7).fill = YELLOW_FILL

    auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    conn.close()
    return len(rows)


# ==================== ERRATA ====================
def build_errata(wb):
    ws = wb.create_sheet("Errata")
    ws.sheet_properties.tabColor = ACCENT_PURPLE
    headers = ["Reporter", "Date", "Code Section", "Description", "Related Proposal", "Confirmed", "Corrected", "Note"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), HEADER_FILL_ERRATA)
    ws.freeze_panes = "A2"

    rows = query("SELECT * FROM errata ORDER BY track, report_date")
    for i, r in enumerate(rows):
        rn = i + 2
        vals = [r["reporter"], r["report_date"], r["code_section"], r["description"],
                r["related_proposal"], "Yes" if r["confirmed"] else "No",
                "Yes" if r["corrected"] else "No", r.get("note")]
        for c_idx, val in enumerate(vals, 1):
            ws.cell(row=rn, column=c_idx, value=val)
        style_data_row(ws, rn, len(headers), alt=(i % 2 == 1))

    auto_width(ws)
    return len(rows)


# ==================== MEETINGS ====================
def build_meetings(wb):
    ws = wb.create_sheet("Meetings")
    ws.sheet_properties.tabColor = ACCENT_ORANGE
    headers = ["Track", "Date", "Time", "Body", "Phase", "Status", "Action Count", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers), HEADER_FILL_MTG)
    ws.freeze_panes = "A2"

    rows = query("SELECT * FROM meetings ORDER BY track DESC, meeting_date")
    for i, r in enumerate(rows):
        rn = i + 2
        vals = [r["track"].capitalize(), r["meeting_date"], r.get("meeting_time"), r["body"],
                r.get("phase"), r["status"], r.get("action_count"), r.get("notes")]
        for c_idx, val in enumerate(vals, 1):
            ws.cell(row=rn, column=c_idx, value=val)
        style_data_row(ws, rn, len(headers), alt=(i % 2 == 1))
        if r["status"] in ("SCHEDULED", "scheduled"):
            ws.cell(row=rn, column=6).fill = YELLOW_FILL

    auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    return len(rows)


# ==================== MAIN ====================
if __name__ == "__main__":
    wb = Workbook()
    build_dashboard(wb)
    com_p = build_proposal_sheet(wb, "commercial", "Commercial", HEADER_FILL_COM)
    res_p = build_proposal_sheet(wb, "residential", "Residential", HEADER_FILL_RES)
    com_ca = build_ca_sheet(wb, "commercial", "Commercial CA", HEADER_FILL_COM)
    res_ca = build_ca_sheet(wb, "residential", "Residential CA", HEADER_FILL_RES)
    com_sa = build_sa_sheet(wb, "commercial", "Commercial SA", HEADER_FILL_COM)
    res_sa = build_sa_sheet(wb, "residential", "Residential SA", HEADER_FILL_RES)
    xo = build_crossovers(wb)
    dq = build_data_quality(wb)
    er = build_errata(wb)
    mt = build_meetings(wb)
    wb.save(OUTPUT)
    print(f"Generated {OUTPUT}")
    print(f"  Dashboard: status, coverage, pending, decisions, ready/awaiting, key dates")
    print(f"  Commercial: {com_p} proposals, {com_ca} CA, {com_sa} SA")
    print(f"  Residential: {res_p} proposals, {res_ca} CA, {res_sa} SA")
    print(f"  Crossovers: {xo} | Data Quality: {dq} | Errata: {er} | Meetings: {mt}")
